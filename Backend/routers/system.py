from functools import lru_cache
from io import BytesIO
import secrets
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from argon2 import PasswordHasher
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from core.security import verify_admin
from database import supabase, supabase_admin
from schemas.core import InternshipCreate, InternshipUpdate, StudentCreate, TeacherCreate, TeacherPasswordReset

router = APIRouter(prefix="/api/system", tags=["System Lookups"])
admin_supabase = supabase_admin or supabase
ph = PasswordHasher()


def _require_admin_supabase():
    if supabase_admin is None:
        raise HTTPException(
            status_code=500,
            detail=(
                "Teacher access requires SUPABASE_SERVICE_ROLE_KEY in Backend/.env "
                "(the Supabase service-role key, not the publishable key)."
            ),
        )
    return supabase_admin


def _friendly_admin_error(exc: Exception) -> str:
    message = str(exc).strip() or "Unknown Supabase admin error."
    lowered = message.lower()
    if "user not allowed" in lowered or "forbidden" in lowered or "permission denied" in lowered:
        return (
            "Supabase rejected teacher creation because the backend is not using a service-role key. "
            "Set SUPABASE_SERVICE_ROLE_KEY in Backend/.env and restart the backend."
        )
    return message


@lru_cache(maxsize=8)
def _get_role_id(role_name: str) -> str | None:
    role_res = admin_supabase.table("Roles").select("id").eq("role_name", role_name).limit(1).execute()
    return role_res.data[0]["id"] if role_res.data else None


@lru_cache(maxsize=1)
def _profiles_has_college_name() -> bool:
    try:
        admin_supabase.table("Profiles").select("college_name").limit(1).execute()
        return True
    except Exception:
        return False


def _relation_to_dict(value: Any) -> dict:
    if isinstance(value, list):
        return value[0] if value else {}
    return value or {}


def _extract_sheet_headers(sheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_idx, cell in enumerate(sheet[1], start=1):
        value = (str(cell.value).strip() if cell.value is not None else "")
        if value:
            headers[value] = col_idx
    return headers


def _attach_teachers(internships: list[dict]) -> list[dict]:
    if not internships:
        return internships

    teacher_role_id = _get_role_id("teacher")
    internship_ids = [row["id"] for row in internships]
    teachers_by_internship: dict[str, list[dict]] = {internship_id: [] for internship_id in internship_ids}

    if teacher_role_id and internship_ids:
        teacher_rows = (
        admin_supabase.table("Profiles")
        .select("id, name, email, internship_id")
        .eq("role_id", teacher_role_id)
        .in_("internship_id", internship_ids)
            .execute()
            .data
            or []
        )
        for teacher in teacher_rows:
            if teacher.get("internship_id") in teachers_by_internship:
                teachers_by_internship[teacher["internship_id"]].append(
                    {"id": teacher["id"], "name": teacher["name"], "email": teacher["email"]}
                )

    for internship in internships:
        teachers = teachers_by_internship.get(internship["id"], [])
        internship["teachers"] = teachers
        internship["teacher_ids"] = [teacher["id"] for teacher in teachers]
    return internships


def _attach_internship_name(teachers: list[dict]) -> list[dict]:
    for teacher in teachers:
        internship = _relation_to_dict(teacher.get("Internships"))
        teacher["internship_name"] = internship.get("name", "Unassigned")
    return teachers


def _sync_internship_teachers(internship_id: str, teacher_ids: list[str] | None):
    if teacher_ids is None:
        return

    teacher_role_id = _get_role_id("teacher")
    if not teacher_role_id:
        raise HTTPException(status_code=500, detail="Teacher role is not configured.")

    current_teachers = (
        supabase.table("Profiles")
        .select("id")
        .eq("role_id", teacher_role_id)
        .eq("internship_id", internship_id)
        .execute()
        .data
        or []
    )
    current_ids = {row["id"] for row in current_teachers}
    requested_ids = {str(teacher_id) for teacher_id in teacher_ids}

    to_clear = list(current_ids - requested_ids)
    if to_clear:
        admin_supabase.table("Profiles").update({"internship_id": None}).in_("id", to_clear).execute()
    if requested_ids:
        admin_supabase.table("Profiles").update({"internship_id": internship_id}).in_("id", list(requested_ids)).execute()


@router.get("/internships/active", status_code=200)
async def get_active_internships():
    try:
        res = (
            admin_supabase.table("Internships")
            .select("id, name, description, is_active")
            .eq("is_active", True)
            .order("created_at", desc=True)
            .execute()
        )
        return res.data
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.get("/roles", status_code=200)
async def get_system_roles():
    try:
        res = admin_supabase.table("Roles").select("id, role_name").execute()
        return res.data
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.get("/teachers", status_code=200)
async def get_system_teachers(auth_data: dict = Depends(verify_admin)):
    try:
        teacher_role_id = _get_role_id("teacher")
        if not teacher_role_id:
            return []

        res = (
            admin_supabase.table("Profiles")
            .select("id, name, email, internship_id, Internships(name)")
            .eq("role_id", teacher_role_id)
            .order("name")
            .execute()
        )
        return _attach_internship_name(res.data or [])
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/teachers", status_code=201)
async def create_system_teacher(
    payload: TeacherCreate,
    auth_data: dict = Depends(verify_admin),
):
    try:
        name = payload.name.strip()
        email = payload.email.strip().lower()
        teacher_role_id = _get_role_id("teacher")
        if not teacher_role_id:
            raise HTTPException(status_code=500, detail="Teacher role is not configured.")

        admin_client = _require_admin_supabase()
        auth_res = admin_client.auth.admin.create_user(
            {
                "email": email,
                "password": payload.password,
                "email_confirm": True,
            }
        )
        created_user = getattr(auth_res, "user", None)
        if not created_user:
            raise HTTPException(status_code=400, detail="Teacher auth account creation failed.")

        profile_payload = {
            "id": created_user.id,
            "name": name,
            "email": email,
            "password": ph.hash(payload.password),
            "role_id": teacher_role_id,
            "internship_id": str(payload.internship_id) if payload.internship_id else None,
        }
        profile_res = admin_supabase.table("Profiles").insert(profile_payload).execute()
        if not profile_res.data:
            try:
                admin_client.auth.admin.delete_user(created_user.id)
            except Exception:
                pass
            raise HTTPException(status_code=400, detail="Teacher profile could not be created.")

        return _attach_internship_name(profile_res.data)[0]
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=_friendly_admin_error(exc)) from exc


@router.post("/students", status_code=201)
async def create_system_student(
    payload: StudentCreate,
    auth_data: dict = Depends(verify_admin),
):
    try:
        name = payload.name.strip()
        email = payload.email.strip().lower()
        student_role_id = _get_role_id("student")
        if not student_role_id:
            raise HTTPException(status_code=500, detail="Student role is not configured.")

        admin_client = _require_admin_supabase()
        auth_res = admin_client.auth.admin.create_user(
            {"email": email, "password": payload.password, "email_confirm": True}
        )
        created_user = getattr(auth_res, "user", None)
        if not created_user:
            raise HTTPException(status_code=400, detail="Student auth account creation failed.")

        profile_payload = {
            "id": created_user.id,
            "name": name,
            "email": email,
            "password": ph.hash(payload.password),
            "role_id": student_role_id,
            "internship_id": str(payload.internship_id) if payload.internship_id else None,
        }
        if _profiles_has_college_name():
            profile_payload["college_name"] = (payload.college_name or "").strip() or None

        profile_res = admin_supabase.table("Profiles").insert(profile_payload).execute()
        if not profile_res.data:
            try:
                admin_client.auth.admin.delete_user(created_user.id)
            except Exception:
                pass
            raise HTTPException(status_code=400, detail="Student profile could not be created.")

        return profile_res.data[0]
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=_friendly_admin_error(exc)) from exc


@router.get("/students/template", status_code=200)
async def download_students_template(auth_data: dict = Depends(verify_admin)):
    try:
        internships = (
            admin_supabase.table("Internships")
            .select("id, name")
            .eq("is_active", True)
            .order("name")
            .execute()
            .data
            or []
        )
        internship_names = [item["name"] for item in internships]

        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.append(["Name", "Email Address", "College Name", "Internship"])
        ws.freeze_panes = "A2"
        ws.append(["Jane Doe", "jane.doe@example.edu", "ABC College", internship_names[0] if internship_names else ""])

        lookup = wb.create_sheet(title="Lookup")
        for idx, name in enumerate(internship_names, start=1):
            lookup.cell(row=idx, column=1, value=name)
        lookup.sheet_state = "hidden"

        if internship_names:
            formula = f"=Lookup!$A$1:$A${len(internship_names)}"
            validation = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(validation)
            validation.add("D2:D5000")

        for col, width in {"A": 28, "B": 34, "C": 28, "D": 30}.items():
            ws.column_dimensions[col].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=students_bulk_template.xlsx"},
        )
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/students/bulk-upload", status_code=200)
async def bulk_upload_students(
    file: UploadFile = File(...),
    auth_data: dict = Depends(verify_admin),
):
    try:
        if not file.filename.lower().endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Please upload a valid .xlsx file.")

        student_role_id = _get_role_id("student")
        if not student_role_id:
            raise HTTPException(status_code=500, detail="Student role is not configured.")

        internships = (
            admin_supabase.table("Internships")
            .select("id, name")
            .eq("is_active", True)
            .execute()
            .data
            or []
        )
        internship_by_name = {item["name"].strip().lower(): item["id"] for item in internships}
        college_enabled = _profiles_has_college_name()
        admin_client = _require_admin_supabase()

        workbook = load_workbook(BytesIO(await file.read()), data_only=True)
        sheet = workbook.active
        headers = _extract_sheet_headers(sheet)
        required = ["Name", "Email Address", "College Name", "Internship"]
        missing = [name for name in required if name not in headers]
        if missing:
            raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}")

        created, updated, skipped = 0, 0, 0
        errors: list[str] = []

        for row_num in range(2, sheet.max_row + 1):
            name = str(sheet.cell(row=row_num, column=headers["Name"]).value or "").strip()
            email = str(sheet.cell(row=row_num, column=headers["Email Address"]).value or "").strip().lower()
            college_name = str(sheet.cell(row=row_num, column=headers["College Name"]).value or "").strip()
            internship_name = str(sheet.cell(row=row_num, column=headers["Internship"]).value or "").strip()

            if not name and not email and not college_name and not internship_name:
                continue
            if not name or not email:
                skipped += 1
                errors.append(f"Row {row_num}: Name and Email Address are required.")
                continue

            internship_id = internship_by_name.get(internship_name.lower()) if internship_name else None
            if internship_name and not internship_id:
                skipped += 1
                errors.append(f"Row {row_num}: Internship '{internship_name}' not found in active internships.")
                continue

            existing = admin_supabase.table("Profiles").select("id").eq("email", email).limit(1).execute().data or []
            payload = {
                "name": name,
                "email": email,
                "internship_id": internship_id,
                "role_id": student_role_id,
            }
            if college_enabled:
                payload["college_name"] = college_name or None

            if existing:
                admin_supabase.table("Profiles").update(payload).eq("id", existing[0]["id"]).execute()
                updated += 1
                continue

            temp_password = f"Temp@{secrets.token_hex(4)}aA1"
            auth_res = admin_client.auth.admin.create_user(
                {"email": email, "password": temp_password, "email_confirm": True}
            )
            created_user = getattr(auth_res, "user", None)
            if not created_user:
                skipped += 1
                errors.append(f"Row {row_num}: Failed to create auth user for {email}.")
                continue

            profile_res = admin_supabase.table("Profiles").insert(
                {**payload, "id": created_user.id, "password": ph.hash(temp_password)}
            ).execute()
            if not profile_res.data:
                try:
                    admin_client.auth.admin.delete_user(created_user.id)
                except Exception:
                    pass
                skipped += 1
                errors.append(f"Row {row_num}: Failed to create profile for {email}.")
                continue
            created += 1

        return {
            "message": "Bulk upload processed.",
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors[:20],
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=_friendly_admin_error(exc)) from exc


@router.put("/teachers/{teacher_id}/password", status_code=200)
async def reset_teacher_password(
    teacher_id: str,
    payload: TeacherPasswordReset,
    auth_data: dict = Depends(verify_admin),
):
    try:
        teacher_role_id = _get_role_id("teacher")
        teacher_res = (
            admin_supabase.table("Profiles")
            .select("id, role_id")
            .eq("id", teacher_id)
            .limit(1)
            .execute()
        )
        if not teacher_res.data:
            raise HTTPException(status_code=404, detail="Teacher not found.")

        teacher = teacher_res.data[0]
        if teacher.get("role_id") != teacher_role_id:
            raise HTTPException(status_code=403, detail="Target user is not a teacher.")

        hashed_password = ph.hash(payload.password)
        update_res = admin_supabase.table("Profiles").update({"password": hashed_password}).eq("id", teacher_id).execute()
        if not update_res.data:
            raise HTTPException(status_code=400, detail="Failed to update teacher password.")

        if supabase_admin is not None:
            try:
                supabase_admin.auth.admin.update_user_by_id(teacher_id, {"password": payload.password})
            except Exception:
                pass

        return {"message": "Teacher password reset successfully."}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/internships", status_code=201)
async def create_internship(
    payload: InternshipCreate,
    auth_data: dict = Depends(verify_admin),
):
    try:
        data = payload.model_dump()
        teacher_ids = [str(teacher_id) for teacher_id in data.pop("teacher_ids", [])]

        res = admin_supabase.table("Internships").insert(data).execute()
        if not res.data:
            raise HTTPException(status_code=400, detail="Internship creation failed.")

        internship = res.data[0]
        _sync_internship_teachers(internship["id"], teacher_ids)
        return _attach_teachers([internship])[0]
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.get("/internships", status_code=200)
async def get_all_internships(auth_data: dict = Depends(verify_admin)):
    try:
        internships = admin_supabase.table("Internships").select("*").order("created_at", desc=True).execute().data or []
        return _attach_teachers(internships)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.put("/internships/{internship_id}", status_code=200)
async def update_internship(
    internship_id: str,
    payload: InternshipUpdate,
    auth_data: dict = Depends(verify_admin),
):
    try:
        update_data = payload.model_dump(exclude_none=True)
        teacher_ids = update_data.pop("teacher_ids", None)

        if update_data:
            res = admin_supabase.table("Internships").update(update_data).eq("id", internship_id).execute()
            if not res.data:
                raise HTTPException(status_code=404, detail="Execution missed natively mapped constraints.")
            internship = res.data[0]
        else:
            existing = admin_supabase.table("Internships").select("*").eq("id", internship_id).limit(1).execute()
            internship = existing.data[0] if existing.data else None

        if not internship:
            raise HTTPException(status_code=404, detail="Internship not found.")

        normalized_teacher_ids = [str(teacher_id) for teacher_id in teacher_ids] if teacher_ids is not None else None
        _sync_internship_teachers(internship_id, normalized_teacher_ids)
        return _attach_teachers([internship])[0]
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.delete("/internships/{internship_id}", status_code=200)
async def delete_internship(
    internship_id: str,
    auth_data: dict = Depends(verify_admin),
):
    try:
        res = admin_supabase.table("Internships").update({"is_active": False}).eq("id", internship_id).execute()
        if not res.data:
            raise HTTPException(status_code=404, detail="Database explicitly failed mapping limits natively.")
        return {"message": "Successfully soft-deleted internship mapping natively rendering array invisible efficiently."}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


from pydantic import BaseModel, Field
from core.logger import logger as app_logger

class LogPayload(BaseModel):
    level: str
    message: str
    user_id: str | None = None
    session_id: str | None = None
    error_code: str | None = None

@router.post("/logs", status_code=200)
async def create_system_log(payload: LogPayload):
    try:
        level_upper = payload.level.upper()
        msg = payload.message
        
        bound_logger = app_logger.bind(
            user_id=payload.user_id,
            session_id=payload.session_id,
            error_code=payload.error_code,
            service="frontend"
        )
        
        if level_upper == "INFO":
            bound_logger.info(msg)
        elif level_upper == "WARNING":
            bound_logger.warning(msg)
        elif level_upper == "ERROR":
            bound_logger.error(msg)
        elif level_upper == "CRITICAL":
            bound_logger.critical(msg)
        else:
            bound_logger.info(msg)
            
        return {"status": "success"}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

